from __future__ import annotations

import logging

logger = logging.getLogger(__name__)

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Mapping
import json
import urllib.parse
import urllib.request
import requests

from ckanapi import RemoteCKAN
from ckanapi.errors import CKANAPIError, NotFound

from ckan_batch.helpers import _to_ckan_payload
from ckan_batch.constants import GCMD_VOCAB_ENDPOINTS, GCMD_BASE_URL


@dataclass
class CreateResult:
    created: List[Dict[str, Any]]          # successful creates (and updates if enabled)
    failed: List[Dict[str, Any]]           # errors with payload + message
    resource_results: List[Dict[str, Any]] = field(default_factory=list)  # per-record resource upload results


def _extract_name_from_ckan_error(err: Any) -> Optional[str]:
    """
    Best-effort helper: tries to extract an existing 'name' from common CKAN error shapes.
    You can remove this if you don't want update behavior.
    """
    try:
        # Sometimes error is a dict like {"name": ["That URL is already in use."]}
        if isinstance(err, dict):
            if "name" in err and isinstance(err["name"], list):
                return None
        # Sometimes error is a string; no reliable extraction without your site's exact message format.
        return None
    except Exception:
        return None


class CKANClient(RemoteCKAN):
    """
    CKAN API client for both:
    - CKAN Action API calls (via RemoteCKAN)
    - Custom CKAN endpoints / blueprint routes (via request_api)

    Example:
        client = CKANClient("https://my-ckan.example", apikey="xxx")

        # Action API
        pkg = client.action.package_show(id="my-dataset")

        # Custom endpoint
        parties = client.get_api("/api/instrument_parties")
    """

    def _build_url(self, path: str) -> str:
        """
        Build an absolute URL from a relative CKAN path.
        Accepts either:
        - '/api/instrument_parties'
        - 'api/instrument_parties'
        - full absolute URL
        """
        if path.startswith("http://") or path.startswith("https://"):
            return path

        base = self.address.rstrip("/") + "/"
        return urllib.parse.urljoin(base, path.lstrip("/"))

    def _get_headers(
        self,
        headers: Optional[Mapping[str, str]] = None,
        apikey: Optional[str] = None,
        json_request: bool = False,
    ) -> Dict[str, str]:
        """
        Build request headers consistent with RemoteCKAN settings.
        """
        final_headers: Dict[str, str] = {
            "User-Agent": self.user_agent,
        }

        key = apikey or self.apikey
        if key:
            # CKAN commonly accepts Authorization; RemoteCKAN itself documents
            # X-CKAN-API-Key for action calls.
            final_headers["Authorization"] = key
            final_headers["X-CKAN-API-Key"] = key

        if json_request:
            final_headers["Accept"] = "application/json"
            final_headers["Content-Type"] = "application/json"

        if headers:
            final_headers.update(headers)

        return final_headers

    @property
    def _http(self) -> requests.Session:
        """
        Reuse the provided session if RemoteCKAN was initialized with one.
        Otherwise create a lightweight session on demand.
        """
        return self.session if self.session is not None else requests.Session()

    def request_api(
        self,
        path: str,
        *,
        method: str = "GET",
        params: Optional[Mapping[str, Any]] = None,
        json: Optional[Any] = None,
        data: Optional[Any] = None,
        files: Optional[Any] = None,
        headers: Optional[Mapping[str, str]] = None,
        apikey: Optional[str] = None,
        timeout: int | float = 60,
        allow_redirects: bool = True,
        raw_response: bool = False,
        **request_kwargs: Any,
    ) -> Any:
        """
        Call a custom CKAN endpoint or any non-Action API route.

        Args:
            path: Relative path like '/api/instrument_parties'
            method: GET, POST, PUT, PATCH, DELETE, etc.
            params: Query string parameters
            json: JSON body
            data: Form body
            files: Files for multipart requests
            headers: Extra headers
            apikey: Override API key for this call
            timeout: Request timeout in seconds
            allow_redirects: Passed to requests
            raw_response: If True, return requests.Response directly
            **request_kwargs: Any extra requests kwargs

        Returns:
            Parsed JSON if possible, otherwise text, unless raw_response=True.
        """
        url = self._build_url(path)
        method = method.upper()

        json_request = json is not None
        req_headers = self._get_headers(
            headers=headers,
            apikey=apikey,
            json_request=json_request,
        )

        response = self._http.request(
            method=method,
            url=url,
            params=params,
            json=json,
            data=data,
            files=files,
            headers=req_headers,
            timeout=timeout,
            allow_redirects=allow_redirects,
            **request_kwargs,
        )

        response.raise_for_status()

        if raw_response:
            return response

        content_type = response.headers.get("Content-Type", "")
        if "application/json" in content_type:
            return response.json()

        return response.text

    # Convenience wrappers
    def get_api(self, path: str, **kwargs: Any) -> Any:
        return self.request_api(path, method="GET", **kwargs)

    def post_api(self, path: str, **kwargs: Any) -> Any:
        return self.request_api(path, method="POST", **kwargs)

    def put_api(self, path: str, **kwargs: Any) -> Any:
        return self.request_api(path, method="PUT", **kwargs)

    def patch_api(self, path: str, **kwargs: Any) -> Any:
        return self.request_api(path, method="PATCH", **kwargs)

    def delete_api(self, path: str, **kwargs: Any) -> Any:
        return self.request_api(path, method="DELETE", **kwargs)


    def create_resources_for_record(
        self,
        package_id: str,
        resources: List[Dict[str, Any]],
        *,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        """
        Upload file resources to an existing CKAN record.
        Each resource dict: {path, name, is_cover, format, description}
        """
        created: List[Dict[str, Any]] = []
        failed: List[Dict[str, Any]] = []

        for res in resources:
            path_str = res.get("path")
            p = Path(path_str) if path_str else None

            if not p or not p.is_file():
                failed.append({"path": path_str, "error": "File not found or not a valid file path"})
                continue

            name = res.get("name") or p.name
            is_cover = res.get("is_cover")
            fmt = res.get("format") or ""
            desc = res.get("description") or ""

            payload = {
                "package_id": package_id,
                "url": "upload",
                "name": name,
                "description": desc,
                "format": fmt,
                "pidinst_is_cover_image": "true" if is_cover else "false",
            }

            if dry_run:
                created.append({"status": "dry_run", "path": str(p), "payload": payload})
                continue

            try:
                with open(p, "rb") as fh:
                    resp = self.action.resource_create(upload=fh, **payload)
                created.append({
                    "id": resp.get("id"),
                    "name": resp.get("name"),
                    "url": resp.get("url"),
                })
            except CKANAPIError as e:
                failed.append({
                    "path": str(p),
                    "error": "CKANAPIError",
                    "ckan_error": getattr(e, "error_dict", None) or str(e),
                })
            except Exception as e:
                failed.append({"path": str(p), "error": f"Unexpected error: {e}"})

        return {"created": created, "failed": failed}

    def create_records(
        self,
        records: List[Dict[str, Any]],
        make_public: bool = False,
        *,
        record_type: str = "instrument",
        dry_run: bool = False,
    ) -> CreateResult:
        """
        Create CKAN records using package_create (or package_update if enabled and exists).

        Assumptions:
          - Your payload dicts match the CKAN scheming fields (e.g. title, owner, manufacturer, model, etc.)
          - CKAN will generate `name` and DOI (if applicable) server-side
        """
        created: List[Dict[str, Any]] = []
        failed: List[Dict[str, Any]] = []
        resource_results: List[Dict[str, Any]] = []

        for i, payload in enumerate(records, start=1):
            # Extract __resources__ before building CKAN payload
            resources = list(payload.get("__resources__") or [])

            # Ensure record_type is set (scheming uses this)
            payload_to_send = dict(payload)
            payload_to_send.pop("__resources__", None)
            payload_to_send["private"] = not make_public
            payload_to_send.setdefault("type", record_type)

            if dry_run:
                rr = self.create_resources_for_record(f"dry_run_{i}", resources, dry_run=True)
                resource_results.append({"index": i, "package_id": None, **rr})
                created.append(
                    {
                        "status": "dry_run",
                        "index": i,
                        "title": payload_to_send.get("title"),
                        "payload": payload_to_send,
                        "resources_dry_run": rr,
                    }
                )
                continue

            try:
                # Create
                payload_to_send = _to_ckan_payload(payload_to_send)  # optional pre-processing if needed
                resp = self.action.package_create(**payload_to_send)
                pkg_id = resp.get("id")
                rr = self.create_resources_for_record(pkg_id, resources, dry_run=dry_run)
                resource_results.append({"index": i, "package_id": pkg_id, **rr})
                created.append(
                    {
                        "status": "created",
                        "index": i,
                        "id": pkg_id,
                        "name": resp.get("name"),
                        "title": resp.get("title"),
                        "doi": resp.get("doi"),  # may be None depending on your site/plugin behavior
                        "response": resp,
                    }
                )

            except CKANAPIError as e:
                # If already exists and update allowed, try update.
                # Note: CKAN typically errors on name collision; but your loader excludes name, so collision is unlikely.
                msg = getattr(e, "error_dict", None) or str(e)

                failed.append(
                    {
                        "index": i,
                        "title": payload_to_send.get("title"),
                        "error": "CKANAPIError",
                        "ckan_error": msg,
                        "payload": payload_to_send,
                    }
                )

            except Exception as e:
                failed.append(
                    {
                        "index": i,
                        "title": payload_to_send.get("title"),
                        "error": f"Unexpected error: {e}",
                        "payload": payload_to_send,
                    }
                )

        return CreateResult(created=created, failed=failed, resource_results=resource_results)


    def delete_record_by_id(self, record_id: str, hard_delete: bool = False) -> bool:
        """
        Delete a record by ID.

        If hard_delete is False, perform a soft delete.
        If hard_delete is True, soft delete first and then purge.

        Returns True if successful, otherwise False.
        """
        try:
            if hard_delete:
                try:
                    self.action.package_delete(id=record_id)
                except CKANAPIError as e:
                    error_text = getattr(e, "error_dict", None) or str(e)
                    text = str(error_text).lower()
                    if "already" not in text or "deleted" not in text:
                        logger.warning("Failed soft delete before purge for record %s: %s", record_id, error_text)
                        return False

                self.action.dataset_purge(id=record_id)
            else:
                self.action.package_delete(id=record_id)

            return True

        except NotFound:
            logger.warning("Record %s not found for deletion", record_id)
            return False
        except CKANAPIError as e:
            logger.warning(
                "CKANAPIError deleting record %s: %s",
                record_id,
                getattr(e, "error_dict", None) or str(e),
            )
            return False
        except Exception:
            logger.exception("Unexpected error deleting record %s", record_id)
            return False


    def delete_all_in_org(
        self,
        owner_org: str = "auscope-org",
        *,
        dry_run: bool = True,
        record_type: Optional[str] = "instrument",
        include_draft: bool = True,
        include_private: bool = True,
        include_public: bool = False,
        hard_delete: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Delete records in an organization with configurable inclusion of draft/private/public.

        IMPORTANT:
        - Draft visibility in `package_search` is controlled by `include_drafts=True`.
        - Private visibility in `package_search` is controlled by `include_private=True`.
        - If hard_delete=True, records are permanently removed using dataset_purge.
        """
        if not (include_private or include_public or include_draft):
            print("Nothing to do: include_draft/include_private/include_public are all False.")
            return []

        owner_org_id = self.get_org_id_by_name(owner_org)
        if owner_org_id is None:
            raise NotFound(f"Organization {owner_org!r} not found.")

        q_parts: List[str] = []
        if record_type is not None:
            q_parts.append(f"type:{record_type}")
            type_label = f"{record_type} "
        else:
            type_label = ""

        q = " AND ".join(q_parts) if q_parts else "*:*"

        fq_parts = [f"owner_org:{owner_org_id}"]

        if include_private and not include_public:
            fq_parts.append("private:true")
        elif include_public and not include_private:
            fq_parts.append("private:false")
        elif not include_public and not include_private:
            print("Nothing to do: both include_private and include_public are False.")
            return []

        if include_draft:
            pass
        else:
            fq_parts.append("state:active")

        fq = " AND ".join(fq_parts)

        include_drafts_flag = bool(include_draft)
        include_private_flag = bool(include_private)

        start = 0
        rows = 100
        to_delete: List[Dict[str, Any]] = []

        while True:
            res = self.action.package_search(
                q=q,
                fq=fq,
                start=start,
                rows=rows,
                include_drafts=include_drafts_flag,
                include_private=include_private_flag,
            )
            results = res.get("results", [])
            if not results:
                break

            for pkg in results:
                to_delete.append(
                    {
                        "id": pkg["id"],
                        "name": pkg["name"],
                        "title": pkg.get("title"),
                        "state": pkg.get("state"),
                        "private": pkg.get("private"),
                    }
                )
            start += rows

        mode = "HARD DELETE" if hard_delete else "SOFT DELETE"
        print(
            f"Found {len(to_delete)} {type_label}record(s) in owner_org={owner_org!r} "
            f"for {mode} "
            f"(draft={include_draft}, private={include_private}, public={include_public})"
        )

        for p in to_delete[:20]:
            vis = "private" if p.get("private") else "public"
            print(f" - {p['name']} ({p['id']}) [{p.get('state')}, {vis}] | {p.get('title')}")
        if len(to_delete) > 20:
            print(f" ... and {len(to_delete) - 20} more")

        if dry_run:
            print("\nDRY RUN: no deletions performed.")
            return to_delete

        deleted = 0
        failed: List[Dict[str, Any]] = []

        for p in to_delete:
            try:
                if hard_delete:
                    self.action.dataset_purge(id=p["id"])
                else:
                    self.action.package_delete(id=p["id"])
                deleted += 1
            except CKANAPIError as e:
                failed.append(
                    {
                        "pkg": p,
                        "error": getattr(e, "error_dict", None) or str(e),
                    }
                )
            except Exception as e:
                failed.append(
                    {
                        "pkg": p,
                        "error": f"Unexpected error: {e}",
                    }
                )

        print(f"\nDeleted: {deleted}")
        if failed:
            print(f"Failed: {len(failed)}")
            for f in failed[:10]:
                print(" -", f)

        return to_delete


    def get_org_id_by_name(self, org_name: str) -> Optional[str]:
        """
        Helper to get organization ID by name.
        Returns None if not found or on error.
        """
        try:
            return self.action.organization_show(id=org_name).get('id')
        except Exception as e:
            print(f"Error fetching organizations: {e}")
            return None


    def get_all(
        self,
        q: str = "*:*",
        fq: Optional[str] = None,
        rows: int = 500,
        include_private: bool = True,
        include_drafts: bool = True,
        verbose: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Returns all packages visible to the API key user.

        Notes:
        - Uses Solr via package_search.
        - You only get what the user can see (private records require permission).
        - `include_private/include_drafts` control CKAN search flags.
        """
        start = 0
        out: List[Dict[str, Any]] = []

        while True:
            res = self.action.package_search(
                q=q,
                fq=fq,
                start=start,
                rows=rows,
                include_private=include_private,
                include_drafts=include_drafts,
            )
            results = res.get("results", [])
            if not results:
                break

            for pkg in results:
                if verbose:
                    out.append(pkg)  # full package dict
                else:
                    out.append(
                        {
                            "id": pkg.get("id"),
                            "name": pkg.get("name"),
                            "title": pkg.get("title"),
                            "type": pkg.get("type"),
                            "state": pkg.get("state"),
                            "owner_org": pkg.get("owner_org"),
                        }
                    )

            start += rows

        return out


    def get_records_by_title(
        self,
        title: str,
        record_type: str | None = None,
        exact_phrase: bool = True,
        rows: int = 100,
        include_private: bool = False,
        include_drafts: bool = False,
    ) -> list[dict]:
        """
        Return a list of records matching the given title.

        Args:
            title: Title value to search for.
            record_type: Optional dataset type filter, e.g. "instrument".
            exact_phrase: If True, search for an exact title match; otherwise do a broader match.
            rows: Maximum number of results to return.
            include_private: Whether to include private datasets.
            include_drafts: Whether to include draft datasets.

        Returns:
            List of matching package dicts. Returns an empty list on error.
        """
        try:
            q = f'title:"{title}"' if exact_phrase else f"title:{title}"
            fq = f"type:{record_type}" if record_type else None

            result = self.action.package_search(
                q=q,
                fq=fq,
                include_private=include_private,
                include_drafts=include_drafts,
                rows=rows,
            )
            return result.get("results", [])

        except NotFound:
            return []
        except CKANAPIError as e:
            print(f"CKANAPIError searching records by title '{title}': {getattr(e, 'error_dict', None) or str(e)}")
            return []
        except Exception as e:
            print(f"Unexpected error searching records by title '{title}': {e}")
            return []

    # ------------------------------------------------------------------ #
    #  Party resolution (cached)                                          #
    # ------------------------------------------------------------------ #
    def _normalize_party_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        out = dict(payload)

        if out.get("website") in ("", None):
            out.pop("website", None)

        if out.get("is_part_of") in ("", None):
            out.pop("is_part_of", None)

        out["type"] = "party"
        return out

    def get_parties_by_name(self) -> Dict[str, Dict[str, Any]]:
        """
        Fetch all party groups and return a dict keyed by lowercased title.
        Cached after first call.
        """
        cache: Optional[Dict[str, Dict[str, Any]]] = getattr(self, "_party_cache", None)
        if cache is not None:
            return cache

        raw = self.action.group_list(
            all_fields=True,
            include_extras=True,
            type="party",
        )

        result: Dict[str, Dict[str, Any]] = {}
        for p in raw:
            title = (p.get("title") or "").strip()
            if not title:
                continue

            roles_raw = p.get("party_role", "")
            if isinstance(roles_raw, list):
                roles = {r.strip().lower() for r in roles_raw if r}
            elif isinstance(roles_raw, str):
                roles = {r.strip().lower() for r in roles_raw.split(",") if r.strip()}
            else:
                roles = set()

            id_type = (p.get("party_identifier_type") or "").strip()
            if id_type.upper() == "ROR":
                identifier = (p.get("party_identifier_ror") or "").strip()
            else:
                identifier = (p.get("party_identifier") or "").strip()

            p_short = {
                "name": p.get("name"),
                "title": title,
                "roles": roles,
                "party_identifier_type": id_type,
                "party_identifier": identifier,
                "party_contact": (p.get("party_contact") or "").strip(),
                # "aliases": p.get("aliases", []),
            }

            result[title.lower()] = p_short
            aliases = [x.strip().lower() for x in p.get("aliases", "").split(',')]
            for al in aliases:
                result[al] = p_short
        self._party_cache = result
        return result

    def create_parties(
        self,
        parties: List[Dict[str, Any]],
        *,
        dry_run: bool = False
    ) -> CreateResult:
        """
        Create CKAN parties using group_create (or group_update if enabled and exists).

        Assumptions:
        - Each payload matches the party scheming group schema.
        - party objects are CKAN groups with type='party'.
        - Parent relationship, if present, is stored in `is_part_of` as the parent party name.
        """
        created: List[Dict[str, Any]] = []
        failed: List[Dict[str, Any]] = []

        for i, payload in enumerate(parties, start=1):
            payload_to_send = dict(payload)
            payload_to_send["type"] = "party"
            payload_to_send = self._normalize_party_payload(payload_to_send)  # optional pre-processing if needed

            if dry_run:
                created.append(
                    {
                        "status": "dry_run",
                        "index": i,
                        "title": payload_to_send.get("title"),
                        "name": payload_to_send.get("name"),
                        "payload": payload_to_send,
                    }
                )
                continue

            try:
                resp = self.action.group_create(**payload_to_send)
                created.append(
                    {
                        "status": "created",
                        "index": i,
                        "id": resp.get("id"),
                        "name": resp.get("name"),
                        "title": resp.get("title"),
                        "response": resp,
                    }
                )

            except CKANAPIError as e:
                msg = getattr(e, "error_dict", None) or str(e)

                failed.append(
                    {
                        "index": i,
                        "title": payload_to_send.get("title"),
                        "name": payload_to_send.get("name"),
                        "error": "CKANAPIError",
                        "ckan_error": msg,
                        "payload": payload_to_send,
                    }
                )

            except Exception as e:
                failed.append(
                    {
                        "index": i,
                        "title": payload_to_send.get("title"),
                        "name": payload_to_send.get("name"),
                        "error": f"Unexpected error: {e}",
                        "payload": payload_to_send,
                    }
                )

        return CreateResult(created=created, failed=failed, resource_results=[])

    def delete_all_parties(
        self,
        *,
        dry_run: bool = True,
        include_only_names: Optional[List[str]] = None,
        hard_delete: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Delete all CKAN groups of type 'party'.

        Args:
            dry_run: If True, only list parties without deleting.
            include_only_names: Optional whitelist of party names to delete.
            hard_delete: If True, permanently remove groups using group_purge.
                If False, perform soft delete using group_delete.

        Returns:
            List of parties that were (or would be) deleted.
        """
        parties = self.action.group_list(
            all_fields=True,
            type="party",
        )

        to_delete: List[Dict[str, Any]] = []
        for grp in parties:
            name = grp.get("name")
            if include_only_names and name not in include_only_names:
                continue

            to_delete.append(
                {
                    "id": grp.get("id"),
                    "name": name,
                    "title": grp.get("title"),
                    "type": grp.get("type"),
                    "state": grp.get("state"),
                }
            )

        mode = "HARD DELETE" if hard_delete else "SOFT DELETE"
        print(f"Found {len(to_delete)} party group(s) for {mode}")
        for g in to_delete[:20]:
            print(
                f" - {g['name']} ({g.get('id')}) | "
                f"{g.get('title')} | state={g.get('state')}"
            )
        if len(to_delete) > 20:
            print(f" ... and {len(to_delete) - 20} more")

        if dry_run:
            print("\nDRY RUN: no deletions performed.")
            return to_delete

        deleted = 0
        failed: List[Dict[str, Any]] = []

        for g in to_delete:
            try:
                if hard_delete:
                    self.action.group_purge(id=g["id"])
                else:
                    self.action.group_delete(id=g["id"])
                deleted += 1
            except CKANAPIError as e:
                failed.append(
                    {
                        "group": g,
                        "error": getattr(e, "error_dict", None) or str(e),
                    }
                )
            except Exception as e:
                failed.append(
                    {
                        "group": g,
                        "error": f"Unexpected error: {e}",
                    }
                )

        print(f"\nDeleted: {deleted}")
        if failed:
            print(f"Failed: {len(failed)}")
            for f in failed[:10]:
                print(" -", f)

        return to_delete


    def get_all_parties(
        self,
        *,
        verbose: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Return all party groups.
        """
        results = self.action.group_list(
            all_fields=True,
            include_extras=True,
            type="party",
        )

        if verbose:
            return results

        return [
            {
                "id": g.get("id"),
                "name": g.get("name"),
                "title": g.get("title"),
                "type": g.get("type"),
                "party_identifier": g.get("party_identifier"),
                "is_part_of": g.get("is_part_of"),
            }
            for g in results
        ]

    # ------------------------------------------------------------------ #
    #  CKAN custom taxonomy resolution (cached)                           #
    # ------------------------------------------------------------------ #

    def get_taxonomy_id_by_name(self, taxonomy_name: str) -> Optional[str]:
        """Return the ID of a CKAN taxonomy by its name. Cached."""
        cache: Optional[List[Dict[str, Any]]] = getattr(self, "_taxonomy_list_cache", None)
        if cache is None:
            cache = self.action.taxonomy_list()
            self._taxonomy_list_cache = cache

        needle = taxonomy_name.strip().lower()
        for t in cache:
            if (t.get("name") or "").strip().lower() == needle:
                return t.get("id")
        return None

    def get_taxonomy_terms(self, taxonomy_id: str) -> List[Dict[str, Any]]:
        """Return terms for a taxonomy by ID. Cached per taxonomy_id."""
        terms_cache: Dict[str, List[Dict[str, Any]]] = getattr(self, "_taxonomy_terms_cache", {})
        if taxonomy_id in terms_cache:
            return terms_cache[taxonomy_id]

        terms = self.action.taxonomy_term_list(id=taxonomy_id)
        terms_cache[taxonomy_id] = terms
        self._taxonomy_terms_cache = terms_cache
        return terms

    def find_taxonomy_term(self, taxonomy_name: str, label: str) -> Optional[Dict[str, Any]]:
        """
        Look up a term by label in a named taxonomy.
        Matches against label, name, and title fields (case-insensitive).
        Returns the term dict or None.
        """
        tid = self.get_taxonomy_id_by_name(taxonomy_name)
        if tid is None:
            return None
        terms = self.get_taxonomy_terms(tid)
        needle = label.strip().lower()
        for t in terms:
            for attr in ("label", "name", "title"):
                val = (t.get(attr) or "").strip()
                if val.lower() == needle:
                    return t
        return None

    # ------------------------------------------------------------------ #
    #  ARDC GCMD vocabulary lookup (cached, via LDA API)                  #
    # ------------------------------------------------------------------ #

    def gcmd_find_term(self, endpoint_key: str, label: str) -> Optional[Dict[str, Any]]:
        """
        Search the ARDC GCMD LDA API for a term by label.
        Returns {"code": <uri>, "label": <prefLabel>} or None.  Cached.
        """
        cache: Dict[Tuple[str, str], Optional[Dict[str, Any]]] = getattr(self, "_gcmd_cache", {})
        norm = label.strip().lower()
        cache_key = (endpoint_key, norm)
        if cache_key in cache:
            return cache[cache_key]

        endpoint = GCMD_VOCAB_ENDPOINTS.get(endpoint_key)
        if not endpoint:
            cache[cache_key] = None
            self._gcmd_cache = cache
            return None

        url = (
            f"{GCMD_BASE_URL}/{endpoint}/concept.json"
            f"?labelcontains={urllib.parse.quote(label.strip())}"
            f"&_pageSize=100"
        )

        try:
            req = urllib.request.Request(url, headers={
                "Accept": "application/json",
                "User-Agent": "ckan-batch/1.0",
            })
            with urllib.request.urlopen(req, timeout=30) as resp:  # noqa: S310
                data = json.loads(resp.read().decode("utf-8"))
        except Exception as exc:
            print(f"[GCMD] HTTP error for {url}: {exc}")
            cache[cache_key] = None
            self._gcmd_cache = cache
            return None

        items = data.get("result", {}).get("items", [])
        for item in items:
            pref = item.get("prefLabel")
            if isinstance(pref, dict):
                pref_val = (pref.get("_value") or "").strip()
            elif isinstance(pref, str):
                pref_val = pref.strip()
            else:
                continue

            if pref_val.lower() == norm:
                result = {"code": item.get("_about", ""), "label": pref_val}
                cache[cache_key] = result
                self._gcmd_cache = cache
                return result

        cache[cache_key] = None
        self._gcmd_cache = cache
        return None


    # ------------------------------------------------------------------ #
    #  Related instrument lookup (by DOI, public only)                    #
    # ------------------------------------------------------------------ #

    def find_public_instrument_by_doi(self, doi: str) -> Optional[Dict[str, Any]]:
        """
        Search for a public, DOI-minted instrument by its DOI value.
        Returns {"id": package_id, "title": title, "doi": doi, "name": slug}
        or None if not found / not public / no minted DOI.
        """
        cache: Dict[str, Optional[Dict[str, Any]]] = getattr(self, "_doi_cache", {})
        norm = doi.strip()
        if norm in cache:
            return cache[norm]

        try:
            results = self.action.package_search(
                q=f'doi:"{norm}"',
                fq="type:instrument",
                include_private=False,
                include_drafts=False,
                rows=5,
            )
        except Exception as exc:
            print(f"[DOI lookup] Search error for {norm}: {exc}")
            cache[norm] = None
            self._doi_cache = cache
            return None

        for pkg in results.get("results", []):
            pkg_doi = (pkg.get("doi") or "").strip()
            if pkg_doi == norm and pkg.get("state") == "active" and not pkg.get("private"):
                result = {
                    "id": pkg["id"],
                    "title": pkg.get("title", ""),
                    "doi": pkg_doi,
                    "name": pkg.get("name", ""),
                }
                cache[norm] = result
                self._doi_cache = cache
                return result

        cache[norm] = None
        self._doi_cache = cache
        return None

    def find_instrument_by_attributes(
        self,
        manufacturer: str,
        model: str,
        alternate_identifier: str,
        visibility: str = "public",
        verbose: bool = False,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[List[Dict[str, str]]]]:
        """
        Search for an instrument by manufacturer name, model name,
        and alternate identifier.

        Args:
            manufacturer: Manufacturer name to match.
            model: Model name to match.
            alternate_identifier: Alternate identifier value to match.
            visibility: One of 'public', 'private', or 'all'.
                - 'public': only public, active records (original behaviour).
                - 'private': only private records.
                - 'all': both public and private records.

        Returns:
        (result_dict, None)         if exactly one match is found
        (None, [summaries])         if multiple matches are found
        (None, None)                if no match is found
        """
        if visibility not in ("public", "private", "all"):
            raise ValueError(f"visibility must be 'public', 'private', or 'all'; got {visibility!r}")

        manuf_norm = manufacturer.strip()
        model_norm = model.strip()
        alt_norm = alternate_identifier.strip()

        include_private = visibility in ("private", "all")
        include_drafts = visibility in ("private", "all")

        try:
            results = self.action.package_search(
                q=(
                    f'manufacturer_name_search:"{manuf_norm}" '
                    f'AND model_name_search:"{model_norm}" '
                    f'AND alternate_identifier_search:"{alt_norm}"'
                ),
                fq="type:instrument",
                include_private=include_private,
                include_drafts=include_drafts,
                rows=100,
            )
        except Exception as exc:
            print(f"[Instrument lookup] Search error: {exc}")
            return None, None

        def _load_list(value):
            if not value:
                return []
            if isinstance(value, str):
                try:
                    value = json.loads(value)
                except (json.JSONDecodeError, TypeError):
                    return []
            return value if isinstance(value, list) else []

        matches: List[Dict[str, Any]] = []

        for pkg in results.get("results", []):
            if pkg.get("state") != "active":
                continue

            # Apply visibility filter
            is_private = pkg.get("private")
            if visibility == "public" and is_private:
                continue
            if visibility == "private" and not is_private:
                continue

            manufacturers = _load_list(pkg.get("manufacturer"))
            manufacturer_match = any(
                (item.get("manufacturer_name") or "").strip() == manuf_norm
                for item in manufacturers
                if isinstance(item, dict)
            )
            if not manufacturer_match:
                continue

            models = _load_list(pkg.get("model"))
            model_match = any(
                (item.get("model_name") or "").strip() == model_norm
                for item in models
                if isinstance(item, dict)
            )
            if not model_match:
                continue

            alternate_ids = _load_list(pkg.get("alternate_identifier_obj"))
            alt_match = any(
                (item.get("alternate_identifier") or "").strip().lower() == alt_norm.lower()
                for item in alternate_ids
                if isinstance(item, dict)
            )
            if not alt_match:
                continue

            matches.append(pkg)

        if len(matches) == 1:
            pkg = matches[0]
            if verbose:
                return pkg, None
            return {
                "id": pkg["id"],
                "title": pkg.get("title", ""),
                "doi": (pkg.get("doi") or "").strip(),
                "name": pkg.get("name", ""),
            }, None

        if len(matches) > 1:
            summaries = [
                {
                    "id": p["id"],
                    "title": p.get("title", ""),
                    "doi": (p.get("doi") or "").strip(),
                }
                for p in matches
            ]
            return None, summaries

        return None, None

    # ------------------------------------------------------------------ #
    #  Export records                                                      #
    # ------------------------------------------------------------------ #

    def export_records(
        self,
        pkg_ids: List[str],
        export_format: str = "Excel",
        output_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Export CKAN records by package IDs to Excel or JSON.

        Args:
            pkg_ids: List of CKAN package IDs to export.
            export_format: 'Excel' or 'JSON'.
            output_path: Optional output file path. If None, auto-generated.

        Returns:
            {"exported": [...], "not_found": [...], "output_path": str}
        """
        if export_format not in ("Excel", "JSON"):
            raise ValueError(f"export_format must be 'Excel' or 'JSON'; got {export_format!r}")

        exported: List[Dict[str, Any]] = []
        not_found: List[str] = []

        for pid in pkg_ids:
            try:
                pkg = self.action.package_show(id=pid)
                exported.append(pkg)
            except NotFound:
                logger.warning("Export: package %s not found", pid)
                not_found.append(pid)
            except CKANAPIError as e:
                logger.warning("Export: CKANAPIError for %s: %s", pid, getattr(e, "error_dict", None) or str(e))
                not_found.append(pid)
            except Exception:
                logger.exception("Export: unexpected error for %s", pid)
                not_found.append(pid)

        logger.info("Exported %d packages, %d not found", len(exported), len(not_found))

        if export_format == "JSON":
            if output_path is None:
                output_path = "export.json"
            with open(output_path, "w", encoding="utf-8") as fh:
                json.dump(exported, fh, ensure_ascii=False, indent=2)

        elif export_format == "Excel":
            if output_path is None:
                output_path = "export.xlsx"

            self._export_to_excel(exported, output_path)
            logger.info("Exported data saved to %s", output_path)

        return {"exported": [p.get("id") for p in exported], "not_found": not_found, "output_path": output_path}

    # ---- Excel export helpers ---- #

    @staticmethod
    def _load_list(value: Any) -> List[Dict[str, Any]]:
        if not value:
            return []
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except (json.JSONDecodeError, TypeError):
                return []
        return value if isinstance(value, list) else []

    def _export_to_excel(self, packages: List[Dict[str, Any]], output_path: str) -> None:
        import openpyxl
        from copy import copy

        template_path = str(Path(__file__).parent / "reader" / "templates" / "PIDINST.xlsx")
        wb = openpyxl.load_workbook(template_path)

        # Ensure both worksheets exist (create Platforms by copying Instruments structure)
        if "Platforms" not in wb.sheetnames:
            src_ws = wb["Instruments"]
            plat_ws = wb.create_sheet("Platforms")
            for row in src_ws.iter_rows(min_row=1, max_row=6, max_col=src_ws.max_column):
                for cell in row:
                    new_cell = plat_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
            for i, dim in src_ws.column_dimensions.items():
                plat_ws.column_dimensions[i].width = dim.width
            # Update title cell
            plat_ws.cell(1, 1, "PIDINST Batch Platform Upload Template (single-sheet, arrays via repeated rows)")

        instruments: List[Dict[str, Any]] = []
        platforms: List[Dict[str, Any]] = []

        for pkg in packages:
            is_platform = str(pkg.get("is_platform", "false")).strip().lower() in ("true", "1", "yes")
            if is_platform:
                platforms.append(pkg)
            else:
                instruments.append(pkg)

        self._write_packages_to_sheet(wb["Instruments"], instruments)
        self._write_packages_to_sheet(wb["Platforms"], platforms)

        wb.save(output_path)

    def _write_packages_to_sheet(self, ws: Any, packages: List[Dict[str, Any]]) -> None:
        """Write package data rows into the given worksheet starting at row 7."""

        def _cell(value: Any) -> Any:
            """Coerce a value to something openpyxl can write."""
            if value is None:
                return ""
            if isinstance(value, list):
                return ", ".join(str(v) for v in value if v is not None)
            return value

        first_data_row = 7
        record_num = 0

        for pkg in packages:
            record_num += 1
            record_label = str(record_num)

            manufacturers  = self._load_list(pkg.get("manufacturer"))
            owners         = self._load_list(pkg.get("owner"))
            models         = self._load_list(pkg.get("model"))
            dates          = self._load_list(pkg.get("date"))
            alt_ids        = self._load_list(pkg.get("alternate_identifier_obj"))
            funders        = self._load_list(pkg.get("funder"))
            related_ids    = self._load_list(pkg.get("related_identifier_obj"))
            instrument_types = self._load_list(pkg.get("instrument_type"))
            measured_vars  = self._load_list(pkg.get("measured_variable"))
            resources      = pkg.get("resources") or []
            if not isinstance(resources, list):
                resources = []

            related_instruments_raw = pkg.get("related_instruments")
            if isinstance(related_instruments_raw, str) and related_instruments_raw.strip():
                try:
                    related_instruments = json.loads(related_instruments_raw)
                except (json.JSONDecodeError, TypeError):
                    related_instruments = []
            elif isinstance(related_instruments_raw, list):
                related_instruments = related_instruments_raw
            else:
                related_instruments = []

            spatial = pkg.get("spatial")
            if isinstance(spatial, str):
                try:
                    spatial = json.loads(spatial)
                except (json.JSONDecodeError, TypeError):
                    spatial = None

            # --- Scalar fields written to the first row only ---
            # user_keywords: list or CSV string → CSV string
            kw_raw = pkg.get("user_keywords", "")
            if isinstance(kw_raw, list):
                user_keywords_str = ", ".join(str(v) for v in kw_raw if v is not None)
            else:
                user_keywords_str = str(kw_raw) if kw_raw else ""

            # Instrument type: split stored list into GCMD (col 16) vs custom (col 17)
            # Use pre-computed instrument_type_gcmd if present; derive custom from list.
            GCMD_HOST = "gcmd.earthdata.nasa.gov"
            it_gcmd_str = pkg.get("instrument_type_gcmd") or ", ".join(
                it.get("instrument_type_name", "") for it in instrument_types
                if GCMD_HOST in (it.get("instrument_type_identifier") or "")
            )
            it_custom_str = ", ".join(
                it.get("instrument_type_name", "") for it in instrument_types
                if GCMD_HOST not in (it.get("instrument_type_identifier") or "")
                and (it.get("instrument_type_identifier") or "")
            )

            # Measured variable: same split
            mv_gcmd_str = pkg.get("measured_variable_gcmd") or ", ".join(
                mv.get("measured_variable_name", "") for mv in measured_vars
                if GCMD_HOST in (mv.get("measured_variable_identifier") or "")
            )
            mv_custom_str = ", ".join(
                mv.get("measured_variable_name", "") for mv in measured_vars
                if GCMD_HOST not in (mv.get("measured_variable_identifier") or "")
                and (mv.get("measured_variable_identifier") or "")
            )

            # Geolocation
            loc_choice = pkg.get("location_choice", "noLocation") or "noLocation"
            lon = lat = min_lng = min_lat = max_lng = max_lat = ""
            if loc_choice == "point" and spatial:
                coords = spatial.get("coordinates") or []
                if len(coords) >= 2:
                    lon, lat = coords[0], coords[1]
            elif loc_choice == "area" and spatial:
                ring = (spatial.get("coordinates") or [[]])[0] or []
                if len(ring) >= 4:
                    min_lng, min_lat = ring[0][0], ring[0][1]
                    max_lng, max_lat = ring[2][0], ring[2][1]

            # Determine total rows needed for this record
            max_rows = max(
                1,
                len(manufacturers),
                len(owners),
                len(models),
                len(dates),
                len(alt_ids),
                len(funders),
                len(related_ids),
                len(related_instruments),
                len(resources),
            )

            for row_idx in range(max_rows):
                r = first_data_row

                # col 1: Record (every row)
                ws.cell(r, 1, record_label)

                # --- First-row-only scalar fields ---
                if row_idx == 0:
                    ws.cell(r,  2, pkg.get("id", ""))                      # PKG_ID
                    ws.cell(r,  3, _cell(pkg.get("title", "")))             # Title
                    ws.cell(r,  4, _cell(pkg.get("instrument_classification", "")))  # Class
                    ws.cell(r, 16, _cell(it_gcmd_str))                     # instrumentTypeGCMD
                    ws.cell(r, 17, _cell(it_custom_str))                   # instrumentTypeCustom
                    ws.cell(r, 27, _cell(mv_gcmd_str))                     # MeasuredVariableGCMD
                    ws.cell(r, 28, _cell(mv_custom_str))                   # MeasuredVariableCustom
                    ws.cell(r, 29, _cell(pkg.get("description", "")))      # Description
                    ws.cell(r, 34, user_keywords_str)                      # UserKeywords
                    ws.cell(r, 35, _cell(pkg.get("credit", "")))           # Credit
                    ws.cell(r, 36, _cell(pkg.get("locality", "")))         # Locality
                    ws.cell(r, 37, loc_choice)                             # Location Type
                    ws.cell(r, 38, _cell(lon))                             # Longitude
                    ws.cell(r, 39, _cell(lat))                             # Latitude
                    ws.cell(r, 40, _cell(min_lng))                         # min_lng
                    ws.cell(r, 41, _cell(min_lat))                         # min_lat
                    ws.cell(r, 42, _cell(max_lng))                         # max_lng
                    ws.cell(r, 43, _cell(max_lat))                         # max_lat
                    ws.cell(r, 44, _cell(pkg.get("epsg_code", "")))        # EPSG

                # col 5: Manufacturer Name
                if row_idx < len(manufacturers):
                    m = manufacturers[row_idx]
                    ws.cell(r, 5, _cell(m.get("manufacturer_name")))

                # col 6-8: Model
                if row_idx < len(models):
                    mdl = models[row_idx]
                    ws.cell(r, 6, _cell(mdl.get("model_name")))
                    ws.cell(r, 7, _cell(mdl.get("model_identifier")))
                    ws.cell(r, 8, _cell(mdl.get("model_identifier_type")))

                # col 9-11: Alternate Identifier
                if row_idx < len(alt_ids):
                    alt = alt_ids[row_idx]
                    ws.cell(r,  9, _cell(alt.get("alternate_identifier")))
                    ws.cell(r, 10, _cell(alt.get("alternate_identifier_type")))
                    ws.cell(r, 11, _cell(alt.get("alternate_identifier_name")))

                # col 12-13: Owner
                if row_idx < len(owners):
                    ow = owners[row_idx]
                    ws.cell(r, 12, _cell(ow.get("owner_name")))
                    ws.cell(r, 13, _cell(ow.get("owner_contact")))

                # col 14-15: Dates
                if row_idx < len(dates):
                    dt = dates[row_idx]
                    ws.cell(r, 14, _cell(dt.get("date_value")))
                    ws.cell(r, 15, _cell(dt.get("date_type")))

                # col 18-22: Related Resources (external)
                if row_idx < len(related_ids):
                    ri = related_ids[row_idx]
                    ws.cell(r, 18, _cell(ri.get("related_identifier")))
                    ws.cell(r, 19, _cell(ri.get("related_identifier_type")))
                    ws.cell(r, 20, _cell(ri.get("related_resource_type")))
                    ws.cell(r, 21, _cell(ri.get("relation_type")))
                    ws.cell(r, 22, _cell(ri.get("related_identifier_name")))

                # col 23: Related Instrument Components (DOI only in export)
                if row_idx < len(related_instruments):
                    rc = related_instruments[row_idx]
                    ws.cell(r, 23, _cell(rc.get("identifier")))
                    # cols 24-26 (Manufacturer/Model/AlternateIdentifier) not stored in relation

                # col 30-33: Funder
                if row_idx < len(funders):
                    fu = funders[row_idx]
                    ws.cell(r, 30, _cell(fu.get("funder_name")))
                    ws.cell(r, 31, _cell(fu.get("award_number")))
                    ws.cell(r, 32, _cell(fu.get("award_uri")))
                    ws.cell(r, 33, _cell(fu.get("award_title")))

                # col 45-49: Resources (use URL as path since we're exporting from CKAN)
                if row_idx < len(resources):
                    res = resources[row_idx]
                    ws.cell(r, 45, _cell(res.get("url")))
                    ws.cell(r, 46, _cell(res.get("name")))
                    is_cover = res.get("pidinst_is_cover_image")
                    if is_cover is True or str(is_cover).lower() in ("true", "1", "yes"):
                        ws.cell(r, 47, "Yes")
                    else:
                        ws.cell(r, 47, "No")
                    ws.cell(r, 48, _cell(res.get("format")))
                    ws.cell(r, 49, _cell(res.get("description")))

                first_data_row += 1

    # ------------------------------------------------------------------ #
    #  Update records                                                    #
    # ------------------------------------------------------------------ #
    def update_records(
        self,
        records: List[Dict[str, Any]],
        convert_to_public: bool = False,
        *,
        dry_run: bool = False,
    ) -> CreateResult:
        """
        Update existing CKAN records.

        Privacy behavior:
        - If convert_to_public=True, force private=False.
        - If convert_to_public=False, preserve the existing package privacy
        unless the incoming payload explicitly includes "private".
        """
        updated: List[Dict[str, Any]] = []
        failed: List[Dict[str, Any]] = []

        for i, payload in enumerate(records, start=1):
            payload_to_send = dict(payload)
            payload_to_send.pop("__resources__", None)

            pkg_id = payload_to_send.pop("pkg_id", None)

            if not pkg_id:
                pkg_id = self._resolve_package_id(payload_to_send, i, failed)
                if pkg_id is None:
                    continue

            try:
                existing = self.action.package_show(id=pkg_id)
            except NotFound:
                failed.append({
                    "index": i,
                    "title": payload_to_send.get("title"),
                    "error": "NotFound",
                    "ckan_error": f"Package {pkg_id!r} not found in database",
                    "payload": payload_to_send,
                })
                continue
            except CKANAPIError as e:
                failed.append({
                    "index": i,
                    "title": payload_to_send.get("title"),
                    "error": "CKANAPIError",
                    "ckan_error": getattr(e, "error_dict", None) or str(e),
                    "payload": payload_to_send,
                })
                continue

            # Privacy handling
            if convert_to_public:
                payload_to_send["private"] = False
            elif "private" not in payload_to_send:
                payload_to_send["private"] = existing.get("private", True)

            payload_to_send["id"] = pkg_id
            payload_to_send.setdefault("type", existing.get("type", "instrument"))

            if dry_run:
                updated.append({
                    "status": "dry_run",
                    "index": i,
                    "id": pkg_id,
                    "title": payload_to_send.get("title"),
                    "payload": payload_to_send,
                })
                continue

            try:
                payload_to_send = _to_ckan_payload(payload_to_send)
                resp = self.action.package_update(**payload_to_send)
                updated.append({
                    "status": "updated",
                    "index": i,
                    "id": resp.get("id"),
                    "name": resp.get("name"),
                    "title": resp.get("title"),
                    "doi": resp.get("doi"),
                    "response": resp,
                })
            except CKANAPIError as e:
                msg = getattr(e, "error_dict", None) or str(e)
                failed.append({
                    "index": i,
                    "title": payload_to_send.get("title"),
                    "error": "CKANAPIError",
                    "ckan_error": msg,
                    "payload": payload_to_send,
                })
            except Exception as e:
                failed.append({
                    "index": i,
                    "title": payload_to_send.get("title"),
                    "error": f"Unexpected error: {e}",
                    "payload": payload_to_send,
                })

        return CreateResult(created=updated, failed=failed, resource_results=[])

    def _resolve_package_id(
        self,
        payload: Dict[str, Any],
        index: int,
        failed: List[Dict[str, Any]],
    ) -> Optional[str]:
        """
        Attempt to resolve a package ID from manufacturer, model, and
        alternate_identifier fields in the payload.
        Appends to *failed* and returns None on error.
        """
        manufacturers = payload.get("manufacturer")
        if isinstance(manufacturers, str):
            try:
                manufacturers = json.loads(manufacturers)
            except (json.JSONDecodeError, TypeError):
                manufacturers = []
        manufacturers = manufacturers if isinstance(manufacturers, list) else []

        models = payload.get("model")
        if isinstance(models, str):
            try:
                models = json.loads(models)
            except (json.JSONDecodeError, TypeError):
                models = []
        models = models if isinstance(models, list) else []

        alt_ids = payload.get("alternate_identifier_obj")
        if isinstance(alt_ids, str):
            try:
                alt_ids = json.loads(alt_ids)
            except (json.JSONDecodeError, TypeError):
                alt_ids = []
        alt_ids = alt_ids if isinstance(alt_ids, list) else []

        manuf_name = (manufacturers[0].get("manufacturer_name") or "").strip() if manufacturers else ""
        model_name = (models[0].get("model_name") or "").strip() if models else ""
        alt_id = (alt_ids[0].get("alternate_identifier") or "").strip() if alt_ids else ""

        if not manuf_name or not model_name or not alt_id:
            failed.append({
                "index": index,
                "title": payload.get("title"),
                "error": "NotFound",
                "ckan_error": (
                    "No pkg_id provided and insufficient attributes for lookup "
                    f"(manufacturer={manuf_name!r}, model={model_name!r}, "
                    f"alternate_identifier={alt_id!r})"
                ),
                "payload": payload,
            })
            return None

        found, duplicates = self.find_instrument_by_attributes(
            manuf_name, model_name, alt_id, visibility="all",
        )

        if found:
            return found["id"]

        if duplicates:
            failed.append({
                "index": index,
                "title": payload.get("title"),
                "error": "MultipleMatches",
                "ckan_error": (
                    f"Multiple packages match manufacturer={manuf_name!r}, "
                    f"model={model_name!r}, alternate_identifier={alt_id!r}: {duplicates}"
                ),
                "payload": payload,
            })
            return None

        failed.append({
            "index": index,
            "title": payload.get("title"),
            "error": "NotFound",
            "ckan_error": (
                f"No package found for manufacturer={manuf_name!r}, "
                f"model={model_name!r}, alternate_identifier={alt_id!r}"
            ),
            "payload": payload,
        })
        return None

    # ------------------------------------------------------------------ #
    #  EPSG code resolution (cached)                                      #
    # ------------------------------------------------------------------ #

    def get_epsg_label(self, code: str) -> str:
        """
        Resolve an EPSG code to its display label (e.g. '4326 - WGS 84').
        Falls back to returning the raw code if the lookup fails.
        Cached per code for the lifetime of this client instance.
        """
        code = str(code).strip()
        if not code:
            return code

        cache: Dict[str, str] = getattr(self, "_epsg_cache", {})
        if code in cache:
            return cache[code]

        url = (
            f"https://apps.epsg.org/api/v1/CoordRefSystem/"
            f"?includeDeprecated=false&pageSize=10&page=0"
            f"&keywords={urllib.parse.quote(code)}"
        )

        try:
            req = urllib.request.Request(url, headers={
                "Accept": "application/json",
                "User-Agent": "ckan-batch/1.0",
            })
            with urllib.request.urlopen(req, timeout=15) as resp:  # noqa: S310
                data = json.loads(resp.read().decode("utf-8"))
        except Exception as exc:
            print(f"[EPSG] Lookup failed for {code}: {exc}")
            cache[code] = code
            self._epsg_cache = cache
            return code

        for item in data.get("Results", []):
            if str(item.get("Code", "")).strip() == code:
                label = f"{code} - {item.get('Name', '')}".strip()
                cache[code] = label
                self._epsg_cache = cache
                return label

        cache[code] = code
        self._epsg_cache = cache
        return code

    def get_taxonomy_list(self):
        return self.action.taxonomy_list()

    def get_taxonomy_term_list(self, taxonomy_id: str):
        return self.action.taxonomy_term_list(id=taxonomy_id)
