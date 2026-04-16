from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import math

import openpyxl

from ckan_batch.constants import COMPOSITE_FIELDS, TAXONOMY_FIELD_MAP
from ckan_batch.helpers import validate_pidinst_date_text, term_to_composite_entry
from ckan_batch.client import CKANClient



# -----------------------------
# Types
# -----------------------------

@dataclass
class MappingResult:
    records: List[Dict[str, Any]]
    errors: List[str]


# -----------------------------
# Helpers
# -----------------------------
def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _clean(v: Any) -> Optional[str]:
    if _is_blank(v):
        return None
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _coerce_bool(v: Any) -> Optional[bool]:
    if _is_blank(v):
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {"true", "t", "yes", "y", "1"}:
        return True
    if s in {"false", "f", "no", "n", "0"}:
        return False
    return None


def _append_unique(lst: List[Dict[str, Any]], item: Dict[str, Any], identity_keys: Tuple[str, ...]) -> None:
    """Append item if its identity tuple is not already present."""
    if not item:
        return
    ident = tuple((_clean(item.get(k)) or "") for k in identity_keys)
    if all(x == "" for x in ident):
        return
    for ex in lst:
        ex_ident = tuple((_clean(ex.get(k)) or "") for k in identity_keys)
        if ident == ex_ident:
            return
    lst.append(item)


def _split_comma_separated(value: Optional[str]) -> Optional[str]:
    """Split comma-separated values, trim each part, and rejoin with ', '."""
    if not value:
        return None
    parts = [part.strip() for part in value.split(",") if part.strip()]
    return ", ".join(parts) if parts else None


def _split_csv_cell(v: Optional[str]) -> List[str]:
    """
    Split a comma-separated cell value into a list of tokens.
    E.g. "a, b ,c" -> ["a","b","c"]
    """
    if not v:
        return []
    parts = [p.strip() for p in str(v).split(",")]
    return [p for p in parts if p]

def _accumulate_csv_field(acc: Dict[str, List[str]], field: str, cell_value: Optional[str]) -> None:
    """
    Accumulate comma-separated values into acc[field] with de-duplication.
    """
    tokens = _split_csv_cell(cell_value)
    if not tokens:
        return
    acc.setdefault(field, [])
    for t in tokens:
        if t not in acc[field]:
            acc[field].append(t)


# -----------------------------
# Party / vocab / geo helpers
# -----------------------------

def _resolve_party_composite(
    party_name: str,
    expected_role: str,
    party_cache: Dict[str, Dict[str, Any]],
    errors: List[str],
    record: str,
    rownum: int,
    *,
    party_id_field: str,
    name_field: str,
    id_field: str,
    id_type_field: str,
) -> Optional[Dict[str, Any]]:
    """Resolve a party by name, validate role, return composite subfield dict."""
    key = party_name.strip().lower()
    party = party_cache.get(key)

    if party is None:
        errors.append(
            f"[Record {record} | Row {rownum}] Party not found: {party_name}"
        )
        return None

    if expected_role.strip().lower() not in party["roles"]:
        errors.append(
            f"[Record {record} | Row {rownum}] Party role mismatch: {party_name} "
            f"expected role {expected_role}"
        )
        return None

    return {
        party_id_field: party["name"],
        name_field: party["title"],
        id_field: party["party_identifier"],
        id_type_field: party["party_identifier_type"],
    }


def _resolve_vocab_items(
    gcmd_raw: Optional[str],
    custom_raw: Optional[str],
    gcmd_endpoint_key: str,
    taxonomy_name: str,
    client: CKANClient,
    errors: List[str],
    record: str,
    *,
    field_name: str,
    field_label: str,
) -> Tuple[List[Dict[str, Any]], Optional[str], Optional[str]]:
    """
    Resolve GCMD + custom taxonomy terms for a controlled-vocab field.
    Returns (items_list, gcmd_codes_csv, gcmd_labels_csv).
    """
    cfg = TAXONOMY_FIELD_MAP[field_name]
    name_field = cfg["name_key"]
    id_field = cfg["identifier_key"]
    id_type_field = cfg["identifier_type_key"]
    items: List[Dict[str, Any]] = []
    gcmd_codes: List[str] = []
    gcmd_labels: List[str] = []

    # --- GCMD terms ---
    if gcmd_raw:
        for token in _split_csv_cell(gcmd_raw):
            found = client.gcmd_find_term(gcmd_endpoint_key, token)
            if found is None:
                errors.append(
                    f"[Record {record}] GCMD term not found ({field_label}): {token}"
                )
            else:
                gcmd_codes.append(found["code"])
                gcmd_labels.append(found["label"])
                items.append({
                    name_field: found["label"],
                    id_field: found["code"],
                    id_type_field: "URL",
                })

    # --- Custom taxonomy terms ---
    if custom_raw:
        for token in _split_csv_cell(custom_raw):
            term = client.find_taxonomy_term(taxonomy_name, token)
            if term is None:
                tax_id = client.get_taxonomy_id_by_name(taxonomy_name)
                lst_term = list(map(lambda x: x["label"], client.get_taxonomy_terms(tax_id)))
                errors.append(
                    f"[Record {record}] Custom taxonomy term not found ({field_label}): {token}"
                )
            else:
                items.append(term_to_composite_entry(term, field_name))

    codes_str = ", ".join(gcmd_codes) if gcmd_codes else None
    labels_str = ", ".join(gcmd_labels) if gcmd_labels else None
    return items, codes_str, labels_str


def _resolve_geolocation(
    rows: List[Dict[str, Any]],
    errors: List[str],
    record: str,
) -> Dict[str, Any]:
    """
    Validate and resolve geolocation fields from grouped rows.
    Returns dict of fields to set on the record.
    """
    loc_types: set = set()
    for row in rows:
        lt = _clean(row.get("GEOLOCATION.Location Type"))
        if lt:
            loc_types.add(lt.strip().lower())

    if not loc_types:
        return {"location_choice": "noLocation"}

    if len(loc_types) > 1:
        errors.append(
            f"[Record {record}] Mixed locationType in record: {loc_types}"
        )
        return {}

    loc_type = loc_types.pop()

    if loc_type == "nolocation":
        return {"location_choice": "noLocation"}

    if loc_type == "point":
        for row in rows:
            lon_raw = _clean(row.get("GEOLOCATION.Longitude"))
            lat_raw = _clean(row.get("GEOLOCATION.Latitude"))
            if lon_raw is not None and lat_raw is not None:
                try:
                    lon = float(lon_raw)
                    lat = float(lat_raw)
                except (ValueError, TypeError):
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Point location requires valid numeric Latitude/Longitude"
                    )
                    return {}

                if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Coordinates out of range: lon={lon}, lat={lat}"
                    )
                    return {}

                fc = {
                    "type": "FeatureCollection",
                    "features": [
                        {
                            "type": "Feature",
                            "properties": {},
                            "geometry": {"type": "Point", "coordinates": [lon, lat]},
                        }
                    ],
                }
                return {
                    "location_choice": "point",
                    "location_data": fc,
                    "spatial": {"type": "Point", "coordinates": [lon, lat]},
                }

        errors.append(
            f"[Record {record}] Point location requires Latitude/Longitude"
        )
        return {}

    if loc_type == "area":
        for row in rows:
            min_lng = _clean(row.get("GEOLOCATION.min_lng"))
            min_lat = _clean(row.get("GEOLOCATION.min_lat"))
            max_lng = _clean(row.get("GEOLOCATION.max_lng"))
            max_lat = _clean(row.get("GEOLOCATION.max_lat"))
            if all(v is not None for v in (min_lng, min_lat, max_lng, max_lat)):
                try:
                    min_lon_f = float(min_lng)
                    min_lat_f = float(min_lat)
                    max_lon_f = float(max_lng)
                    max_lat_f = float(max_lat)
                except (ValueError, TypeError):
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Area location requires valid numeric bounding box coordinates"
                    )
                    return {}

                if min_lon_f > max_lon_f or min_lat_f > max_lat_f:
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Area bounding box invalid: min > max"
                    )
                    return {}

                if not (-90 <= min_lat_f <= 90) or not (-90 <= max_lat_f <= 90):
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Latitude out of range (-90 to 90)"
                    )
                    return {}

                if not (-180 <= min_lon_f <= 180) or not (-180 <= max_lon_f <= 180):
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Longitude out of range (-180 to 180)"
                    )
                    return {}

                coords = [
                    [min_lon_f, min_lat_f],
                    [max_lon_f, min_lat_f],
                    [max_lon_f, max_lat_f],
                    [min_lon_f, max_lat_f],
                    [min_lon_f, min_lat_f],
                ]
                fc = {
                    "type": "FeatureCollection",
                    "features": [
                        {
                            "type": "Feature",
                            "properties": {},
                            "geometry": {"type": "Polygon", "coordinates": [coords]},
                        }
                    ],
                }
                return {
                    "location_choice": "area",
                    "location_data": fc,
                    "spatial": {"type": "Polygon", "coordinates": [coords]},
                }

        errors.append(
            f"[Record {record}] Area location requires min/max bounding box coordinates"
        )
        return {}

    errors.append(
        f"[Record {record}] Unknown locationType: {loc_type}"
    )
    return {}


# -----------------------------
# Template header parsing
# -----------------------------
def _forward_fill(values: List[Optional[str]]) -> List[Optional[str]]:
    out: List[Optional[str]] = []
    cur: Optional[str] = None
    for v in values:
        if v is not None and str(v).strip() != "":
            cur = str(v).strip()
        out.append(cur)
    return out


def _norm_header(label: Optional[str]) -> Optional[str]:
    if label is None:
        return None
    s = str(label).strip()
    return s if s else None


def _strip_required_star(col: str) -> Tuple[str, bool]:
    col = col.strip()
    if col.endswith("*"):
        return col[:-1].strip(), True
    return col, False


def _build_column_keys(ws, header_row: int = 5, section_row: int = 3, group_row: int = 4) -> Tuple[List[str], Dict[str, bool]]:
    """
    Builds stable, unique column keys using the multi-row header layout:
      - section_row (row 3): e.g. OWNER, DATES, etc (sparse)
      - group_row   (row 4): e.g. MODEL, MANUFACTURER, etc (sparse)
      - header_row  (row 5): actual column names (often repeated: Name*, Identifier, etc)

    Output keys look like:
      - "TITLE.Name"
      - "MANUFACTURER.IdentifierType"
      - "OWNER.Contact"
      - "RELATED_RESOURCES_REGISTERED.Serial Number"
      - "GEOLOCATION.Latitude"
    """
    max_col = ws.max_column

    sec = [_norm_header(ws.cell(section_row, c).value) for c in range(1, max_col + 1)]
    grp = [_norm_header(ws.cell(group_row, c).value) for c in range(1, max_col + 1)]
    hdr_raw = [_norm_header(ws.cell(header_row, c).value) for c in range(1, max_col + 1)]

    sec_ff = _forward_fill(sec)
    grp_ff = _forward_fill(grp)

    required_map: Dict[str, bool] = {}
    keys: List[str] = []

    for i in range(max_col):
        h = hdr_raw[i]
        if h is None:
            keys.append(f"__EMPTY__{i+1}")
            required_map[keys[-1]] = False
            continue

        h2, is_req = _strip_required_star(h)

        # Special-case: section rows in your template contain long text for the related blocks
        # Normalize those into short identifiers.
        sec_label = sec_ff[i] or ""
        if sec_label.startswith("RELATED_RESOURCES") or sec_label.startswith("RELATED RESOURCES"):
            sec_label = "RELATED_RESOURCES"
        elif sec_label.startswith("RELATED_INSTRUMENT_COMPONENTS") or sec_label.startswith("RELATED INSTRUMENT COMPONENTS"):
            sec_label = "RELATED_INSTRUMENT_COMPONENTS"
        elif sec_label.startswith("RESOURCES"):
            sec_label = "RESOURCES"

        grp_label = grp_ff[i] or ""
        # Prefer group label (row 4) when present, else section label (row 3)
        prefix = grp_label or sec_label or "FIELD"

        key = f"{prefix}.{h2}"
        keys.append(key)
        required_map[key] = is_req

    return keys, required_map

# -----------------------------
# Main mapper
# -----------------------------
def read_pidinst_template(
    excel_path: str,
    client: CKANClient,
    sheet_name: str = "Instruments",
    record_col_key: str = "FIELD.Record",  # derived key for Record* column
    org_own: str = "auscope-org",  # default owner organization (CKAN org name, not required in your template)
) -> MappingResult:
    """
    Reads your adjusted PIDINST template and maps it to CKAN record payload dicts.

    Grouping:
      - uses Record* as the grouping key (first row in a group holds base required fields)

    Repeating composites supported:
      - manufacturer, owner, model, date, alternate_identifier_obj, funder,
        related_identifier_obj, related_instruments

    is_platform is derived from sheet_name:
      - "Instruments" -> False
      - "Platforms"   -> True
    """
    _IS_PLATFORM_SHEETS = {"Platforms": "true"}
    sheet_is_platform = _IS_PLATFORM_SHEETS.get(sheet_name, "false")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    col_keys, required_cols = _build_column_keys(ws, header_row=5, section_row=3, group_row=4)

    # Data starts at row 7 in this template:
    # row 1 title, row 2 note, row 3/4 groups, row 5 headers, row 6 help, row 7+ data
    first_data_row = 7

    # Build row dicts with our stable keys
    rows: List[Dict[str, Any]] = []
    for r in range(first_data_row, ws.max_row + 1):
        row: Dict[str, Any] = {"__rownum__": r}
        empty = True
        for c, key in enumerate(col_keys, start=1):
            val = ws.cell(r, c).value
            if not _is_blank(val):
                empty = False
            row[key] = val
        if not empty:
            rows.append(row)

    errors: List[str] = []
    records: List[Dict[str, Any]] = []

    # Group by Record*
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        rec = _clean(row.get(record_col_key))
        if not rec:
            # If Record* is empty, ignore row (or flag – your choice)
            errors.append(f"[Row {row['__rownum__']}] Missing Record* (used for grouping).")
            continue
        groups.setdefault(rec, []).append(row)

    # Resolve party cache once
    try:
        party_cache = client.get_parties_by_name()
    except Exception as exc:
        errors.append(f"Failed to fetch party list: {exc}")
        return MappingResult(records=[], errors=errors)

    for record, grp in groups.items():
        ds: Dict[str, Any] = {}

        # Repeating composites (lists)
        for k in COMPOSITE_FIELDS:
            ds[k] = []

        # Related instrument components (JSON list for the picker field)
        ds["related_instruments"] = []

        # Tag-like accumulators
        tag_acc: Dict[str, List[str]] = {}

        # Vocab field token accumulators (instrument type + measured variable)
        it_gcmd_tokens: List[str] = []
        it_custom_tokens: List[str] = []
        mv_gcmd_tokens: List[str] = []
        mv_custom_tokens: List[str] = []


        # Validate required columns on the first row (your “first row has required info” rule)
        first = grp[0]
        missing_required_headers: List[str] = []
        for k, is_req in required_cols.items():
            if not is_req:
                continue
            if k == record_col_key:
                continue
            if _is_blank(first.get(k)):
                missing_required_headers.append(k)

        if missing_required_headers:
            errors.append(
                f"[Record {record}] Missing required values on first row: {', '.join(missing_required_headers)}"
            )
            continue

        for row in grp:
            # Base scalar fields
            ds["owner_org"] = org_own  # CKAN owner organization (not required in your template, but if present on any row in the group, use it)
            title = _clean(row.get("TITLE.Name"))
            if title and _is_blank(ds.get("title")):
                ds["title"] = title  # CKAN title

            instrument_class = _clean(row.get("INSTRUMENT_CLASS.Class"))
            if instrument_class and _is_blank(ds.get("instrument_classification")):
                ds["instrument_classification"] = instrument_class  # CKAN instrument class

            desc = _clean(row.get("OTHER.Description"))
            if desc and _is_blank(ds.get("description")):
                ds["description"] = desc

            locality = _clean(row.get("GEOLOCATION.Locality"))
            if locality and _is_blank(ds.get("locality")):
                ds["locality"] = locality

            epsg = _clean(row.get("GEOLOCATION.EPSG"))
            if epsg and _is_blank(ds.get("epsg_code")):
                ds["epsg_code"] = epsg
                ds["epsg"] = client.get_epsg_label(epsg)

            # (Geolocation is resolved per-record after all rows – see below)

            # Comma-separated tag fields (read from first row with value)
            if _is_blank(ds.get("user_keywords")):
                user_kw = _split_comma_separated(_clean(row.get("OTHER.UserKeywords")))
                if user_kw:
                    ds["user_keywords"] = user_kw

            # Tags / multi-token fields (accumulate across rows)
            _accumulate_csv_field(tag_acc, "user_keywords", _clean(row.get("OTHER.UserKeywords")))

            # Instrument type tokens (accumulate across rows)
            for t in _split_csv_cell(_clean(row.get("INSTRUMENT (RESOURCE) TYPE.instrumentTypeGCMD"))):
                if t not in it_gcmd_tokens:
                    it_gcmd_tokens.append(t)
            for t in _split_csv_cell(_clean(row.get("INSTRUMENT (RESOURCE) TYPE.instrumentTypeCustom"))):
                if t not in it_custom_tokens:
                    it_custom_tokens.append(t)

            # Measured variable tokens (accumulate across rows)
            for t in _split_csv_cell(_clean(row.get("OTHER.MeasuredVariableGCMD"))):
                if t not in mv_gcmd_tokens:
                    mv_gcmd_tokens.append(t)
            for t in _split_csv_cell(_clean(row.get("OTHER.MeasuredVariableCustom"))):
                if t not in mv_custom_tokens:
                    mv_custom_tokens.append(t)

            credit = _clean(row.get("OTHER.Credit"))
            if credit and _is_blank(ds.get("credit")):
                ds["credit"] = credit

            # Manufacturer composite (repeating) – resolved via party registry
            manuf_name_raw = _clean(row.get("MANUFACTURER.Name"))
            if manuf_name_raw:
                m = _resolve_party_composite(
                    manuf_name_raw, "Manufacturer", party_cache, errors,
                    record, row["__rownum__"],
                    party_id_field="manufacturer_party_id",
                    name_field="manufacturer_name",
                    id_field="manufacturer_identifier",
                    id_type_field="manufacturer_identifier_type",
                )
                if m:
                    _append_unique(ds["manufacturer"], m, identity_keys=("manufacturer_name",))

            # Owner composite (repeating) – resolved via party registry
            owner_name_raw = _clean(row.get("OWNER.Name"))
            owner_contact_raw = _clean(row.get("OWNER.Contact"))
            if owner_name_raw:
                o = _resolve_party_composite(
                    owner_name_raw, "Owner", party_cache, errors,
                    record, row["__rownum__"],
                    party_id_field="owner_party_id",
                    name_field="owner_name",
                    id_field="owner_identifier",
                    id_type_field="owner_identifier_type",
                )
                if o:
                    contact = owner_contact_raw
                    if not contact:
                        party_entry = party_cache.get(owner_name_raw.strip().lower())
                        if party_entry:
                            contact = party_entry.get("party_contact")
                    o["owner_contact"] = contact or ""
                    _append_unique(ds["owner"], o, identity_keys=("owner_name",))

            # Model composite (repeating)
            model = {
                "model_name": _clean(row.get("MODEL.Name")),
                "model_identifier": _clean(row.get("MODEL.Identifier")),
                "model_identifier_type": _clean(row.get("MODEL.IdentifierType")),
            }
            if model.get("model_name"):
                _append_unique(ds["model"], model, identity_keys=("model_name", "model_identifier"))

            # Date composite (repeating)
            dtype = _clean(row.get("DATES.dateType"))
            if dtype and dtype.lower() == "period of activity":
                dtype = "Coverage"
            raw_dval = row.get("DATES.Date")
            try:
                dval = validate_pidinst_date_text(raw_dval, date_type=dtype)
            except ValueError as exc:
                errors.append(f"[Record {record} | Row {row['__rownum__']}] {exc}")
                dval = None

            if dval or dtype:
                date_obj = {"date_value": dval, "date_type": dtype}
                _append_unique(ds["date"], date_obj, identity_keys=("date_value", "date_type"))
            # Alternate Identifier composite (repeating)
            alt = {
                "alternate_identifier": _clean(row.get("ALTERNATE IDENTIFIER.Id")),
                "alternate_identifier_type": _clean(row.get("ALTERNATE IDENTIFIER.IdType")),
                "alternate_identifier_name": _clean(row.get("ALTERNATE IDENTIFIER.Name")),
            }
            if alt.get("alternate_identifier") or alt.get("alternate_identifier_type"):
                _append_unique(
                    ds["alternate_identifier_obj"],
                    alt,
                    identity_keys=("alternate_identifier_type", "alternate_identifier"),
                )

            # Resources / attachments
            res_path = _clean(row.get("ATTACHMENTS.Path"))
            if res_path:
                # Expand ~ to the user's home directory
                res_path = str(Path(res_path).expanduser())
                res_name = _clean(row.get("ATTACHMENTS.Name"))
                res_is_cover = _coerce_bool(_clean(row.get("ATTACHMENTS.IsCover")))
                res_fmt = _clean(row.get("ATTACHMENTS.Format"))
                res_desc = _clean(row.get("ATTACHMENTS.Description"))
                ds.setdefault("__resources__", []).append({
                    "path": res_path,
                    "name": res_name,
                    "is_cover": res_is_cover,
                    "format": res_fmt,
                    "description": res_desc,
                })

            # Funder composite (repeating) – resolved via party registry
            funder_name_raw = _clean(row.get("FUNDER.Name"))
            if funder_name_raw:
                f = _resolve_party_composite(
                    funder_name_raw, "Funder", party_cache, errors,
                    record, row["__rownum__"],
                    party_id_field="funder_party_id",
                    name_field="funder_name",
                    id_field="funder_identifier",
                    id_type_field="funder_identifier_type",
                )
                if f:
                    f["award_number"] = _clean(row.get("FUNDER.AwardNumber"))
                    f["award_uri"] = _clean(row.get("FUNDER.AwardURI"))
                    f["award_title"] = _clean(row.get("FUNDER.AwardTitle"))
                    _append_unique(ds["funder"], f, identity_keys=("funder_name", "award_number"))

            # -----------------------------
            # Related resources (external)
            # -----------------------------
            ext_id = _clean(row.get("RELATED_RESOURCES.Id"))
            ext_id_type = _clean(row.get("RELATED_RESOURCES.IdType"))
            ext_res_type = _clean(row.get("RELATED_RESOURCES.ResourceType"))
            ext_rel = _clean(row.get("RELATED_RESOURCES.Relationship"))
            ext_name = _clean(row.get("RELATED_RESOURCES.IdentifierName"))

            if ext_id or ext_res_type or ext_rel:
                rel_obj = {
                    "related_identifier": ext_id,
                    "related_identifier_type": ext_id_type,
                    "related_resource_type": ext_res_type,
                    "relation_type": ext_rel,
                    "related_identifier_name": ext_name,
                }
                if rel_obj.get("related_identifier"):
                    _append_unique(
                        ds["related_identifier_obj"],
                        rel_obj,
                        identity_keys=("related_identifier_type", "related_identifier", "relation_type"),
                    )

            # -----------------------------
            # Related instrument components
            # -----------------------------
            comp_id = _clean(row.get("RELATED_INSTRUMENT_COMPONENTS.Id"))
            comp_manf = _clean(row.get("RELATED_INSTRUMENT_COMPONENTS.Manufacturer"))
            comp_model = _clean(row.get("RELATED_INSTRUMENT_COMPONENTS.Model"))
            comp_alt = _clean(row.get("RELATED_INSTRUMENT_COMPONENTS.AlternateIdentifier"))

            if comp_id:
                # Look up instrument by DOI - must be public with minted DOI
                found = client.find_public_instrument_by_doi(comp_id)
                if found:
                    _append_unique(
                        ds["related_instruments"],
                        {
                            "package_id": found["id"],
                            "identifier": found["doi"],
                            "identifier_type": "DOI",
                            "label": found["title"],
                            "relation_type": "HasPart",
                        },
                        identity_keys=("package_id",),
                    )
                else:
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Related instrument component not found as a public instrument with minted DOI: {comp_id}"
                    )
            elif comp_manf and comp_model and comp_alt:
                # Search by Manufacturer + Model + AlternateIdentifier combo
                found, duplicates = client.find_public_instrument_by_attributes(
                    comp_manf, comp_model, comp_alt,
                )
                if found:
                    _append_unique(
                        ds["related_instruments"],
                        {
                            "package_id": found["id"],
                            "identifier": found["doi"],
                            "identifier_type": "DOI",
                            "label": found["title"],
                            "relation_type": "HasPart",
                        },
                        identity_keys=("package_id",),
                    )
                elif duplicates:
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"Multiple public instruments match Manufacturer={comp_manf}, "
                        f"Model={comp_model}, AlternateIdentifier={comp_alt}: {duplicates}"
                    )
                else:
                    errors.append(
                        f"[Record {record} | Row {row['__rownum__']}] "
                        f"No public instrument found for Manufacturer={comp_manf}, "
                        f"Model={comp_model}, AlternateIdentifier={comp_alt}"
                    )
            elif comp_manf or comp_model or comp_alt:
                errors.append(
                    f"[Record {record} | Row {row['__rownum__']}] "
                    f"Related instrument component requires either a DOI in the Id column "
                    f"or all three: Manufacturer, Model, and AlternateIdentifier."
                )

        # Apply accumulated tag fields (join)
        for k, vals in tag_acc.items():
            if vals:
                ds[k] = ", ".join(vals)

        # ----- Instrument type resolution (GCMD + custom taxonomy) -----
        is_platform = sheet_is_platform == "true"
        gcmd_ep = "platforms" if is_platform else "instruments"
        tax_name = "platforms" if is_platform else "instruments"
        if it_gcmd_tokens or it_custom_tokens:
            gcmd_joined = ", ".join(it_gcmd_tokens) if it_gcmd_tokens else None
            custom_joined = ", ".join(it_custom_tokens) if it_custom_tokens else None
            it_items, it_gcmd_codes, it_gcmd_labels = _resolve_vocab_items(
                gcmd_joined, custom_joined,
                gcmd_ep, tax_name, client, errors, record,
                field_name="instrument_type",
                field_label="instrument type",
            )
            if it_items:
                ds["instrument_type"] = it_items
            if it_gcmd_codes:
                ds["instrument_type_gcmd_code"] = it_gcmd_codes
            if it_gcmd_labels:
                ds["instrument_type_gcmd"] = it_gcmd_labels

        # ----- Measured variable resolution (GCMD + custom taxonomy) -----
        if mv_gcmd_tokens or mv_custom_tokens:
            gcmd_joined = ", ".join(mv_gcmd_tokens) if mv_gcmd_tokens else None
            custom_joined = ", ".join(mv_custom_tokens) if mv_custom_tokens else None
            mv_items, mv_gcmd_codes, mv_gcmd_labels = _resolve_vocab_items(
                gcmd_joined, custom_joined,
                "measured_variables", "measured-variables", client, errors, record,
                field_name="measured_variable",
                field_label="measured variable",
            )
            if mv_items:
                ds["measured_variable"] = mv_items
            if mv_gcmd_codes:
                ds["measured_variable_gcmd_code"] = mv_gcmd_codes
            if mv_gcmd_labels:
                ds["measured_variable_gcmd"] = mv_gcmd_labels

        # ----- Geolocation resolution (locationType-based) -----
        geo = _resolve_geolocation(grp, errors, record)
        ds.update(geo)

        # Clean empty repeating composites
        for k in COMPOSITE_FIELDS:
            if not ds.get(k):
                ds.pop(k, None)

        # Serialize related_instruments to JSON (the validator expects a JSON string)
        if ds.get("related_instruments"):
            ds["related_instruments"] = json.dumps(ds["related_instruments"], ensure_ascii=False)
        else:
            ds.pop("related_instruments", None)

        # Clean empty __resources__
        if not ds.get("__resources__"):
            ds.pop("__resources__", None)

        # Minimal required checks (based on your template)
        if _is_blank(ds.get("title")):
            errors.append(f"[Record {record}] Missing TITLE.Name (maps to CKAN title).")
            continue
        if not ds.get("manufacturer"):
            errors.append(f"[Record {record}] Missing at least one manufacturer (MANUFACTURER.Name).")
            continue
        if not ds.get("owner"):
            errors.append(f"[Record {record}] Missing at least one owner (OWNER.Name/Contact/Relationship).")
            continue
        if not ds.get("model"):
            errors.append(f"[Record {record}] Missing at least one model (MODEL.Name).")
            continue

        ds["is_platform"] = sheet_is_platform
        records.append(ds)

    return MappingResult(records=records, errors=errors)

